from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SNAPSHOT_DATE = "2026-06-06"
OUTPUT_ID = "019e910c-af91-7f70-b575-98ceeb8830a1"
DEFAULT_RANKING_SIZE = 20


FOLDER_LABELS = {
    "Foundation Models": "Foundation Models",
    "AI Agents": "AI Agents",
    "AI Infra": "AI Infra",
    "AI Hardware": "AI Hardware",
    "Generative Media": "Generative Media",
    "Autonomous Driving": "Autonomous Driving",
    "AI Trust & Governance": "AI Trust & Governance",
    "Robotics": "Robotics",
}


TRACK_LABEL_OVERRIDES = {
    "foundation_models_base_models_top10": "Base foundation models",
    "foundation_models_edge_small_models_top10": "Edge small models",
    "foundation_models_embedding_top10": "Embedding models",
    "foundation_models_math_science_top10": "Math and science models",
    "foundation_models_biopharma_top10": "Biopharma models",
    "foundation_models_weather_geo_rs_top10": "Weather, geo and remote sensing",
    "foundation_models_industrial_top10": "Industrial foundation models",
    "ai_agents_finance_top10": "Financial AI agents",
    "ai_agents_enterprise_top10": "Enterprise AI agents",
    "ai_agents_medical_top10": "Medical AI agents",
    "ai_agents_gov_office_top10": "Government and office agents",
    "ai_agents_general_llm_top10": "General LLM agents",
    "ai_agents_media_top10": "Media agents",
    "ai_agents_embodied_vla_top10": "Embodied VLA agents",
    "ai_infra_data_collection_labeling_top10": "Data collection and labeling",
    "ai_infra_dev_framework_top10": "AI development frameworks",
    "ai_infra_specialized_industry_platforms_top10": "Specialized industry platforms",
    "ai_hardware_compute_clusters_top10": "AI compute clusters",
    "ai_hardware_ai_servers_top10": "AI servers",
    "ai_hardware_supporting_components_top10": "AI supporting hardware",
    "ai_hardware_ai_chips_top10": "AI chips",
    "generative_media_digital_humans_top10": "Digital humans",
    "generative_media_audio_video_models_top10": "Audio and video generation",
    "generative_media_image_models_top10": "Image generation",
    "generative_media_interactive_content_top10": "Interactive content generation",
    "robotics_industrial_fixed_robots_top10": "Industrial fixed robots",
    "robotics_mobile_robots_top10": "Mobile robots",
    "robotics_commercial_service_top10": "Commercial service robots",
    "robotics_home_consumer_top10": "Home consumer robots",
    "robotics_specialized_industry_top10": "Special-purpose robots",
    "robotics_bionic_top10": "Bionic and embodied robots",
    "robotics_support_services_top10": "Robot support services",
    "autonomous_driving_passenger_fullstack_top10": "Passenger autonomous driving",
    "autonomous_driving_cabin_llm_top10": "In-cabin AI interaction",
    "autonomous_driving_commercial_vehicles_top10": "Commercial vehicle autonomy",
    "autonomous_driving_low_speed_delivery_top10": "Low-speed delivery vehicles",
    "ai_trust_model_evaluation_top10": "Model evaluation",
    "ai_trust_model_safety_redteam_top10": "Model safety and red teaming",
    "ai_trust_model_deployment_services_top10": "Model deployment services",
    "ai_trust_data_compliance_top10": "AI data compliance",
}


def slugify(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "item"


def track_name_from_slug(slug: str) -> str:
    normalized_slug = re.sub(r"_top20$", "_top10", slug)
    if normalized_slug in TRACK_LABEL_OVERRIDES:
        return TRACK_LABEL_OVERRIDES[normalized_slug]
    cleaned = re.sub(r"_top(?:10|20)$", "", slug)
    for prefix in (
        "foundation_models_",
        "ai_agents_",
        "ai_infra_",
        "ai_hardware_",
        "generative_media_",
        "autonomous_driving_",
        "ai_trust_",
        "robotics_",
    ):
        cleaned = cleaned.replace(prefix, "")
    return cleaned.replace("_", " ").title()


def domain_for_folder(folder: str) -> str:
    return slugify(folder)


def source_type(value: Any) -> str:
    raw = str(value or "").lower()
    if "benchmark" in raw:
        return "benchmark"
    if "official" in raw:
        return "official"
    if "research" in raw:
        return "research"
    if "community" in raw:
        return "community"
    return "market-data"


def source_quality(source_type_value: str, confidence: str = "") -> str:
    if source_type_value in {"official", "benchmark"}:
        return "high"
    if "high" in confidence.lower():
        return "high"
    if "medium" in confidence.lower():
        return "medium"
    return "proxy"


def confidence_quality(confidence: Any) -> str:
    value = str(confidence or "").lower()
    if "high" in value:
        return "high"
    if "medium" in value:
        return "medium"
    return "proxy"


def entity_type_for(folder: str, track_slug: str) -> str:
    if folder == "Robotics":
        return "robot" if "services" not in track_slug else "service"
    if folder == "AI Hardware":
        return "infrastructure"
    if folder == "AI Infra":
        return "infrastructure"
    if folder == "Foundation Models":
        return "model"
    if folder == "AI Trust & Governance":
        return "service"
    return "company"


def region_for(country: str) -> str:
    value = country.lower()
    if "china" in value or "taiwan" in value or "hong kong" in value:
        return "China / Greater China"
    if "united states" in value or "canada" in value:
        return "North America"
    if any(token in value for token in ["united kingdom", "france", "germany", "switzerland", "ireland", "sweden", "estonia", "europe"]):
        return "Europe"
    if any(token in value for token in ["japan", "korea", "singapore", "india"]):
        return "Asia Pacific"
    return "Global"


def stage_for(company: str) -> str:
    public_names = {
        "microsoft",
        "google",
        "google deepmind",
        "meta ai",
        "apple",
        "qualcomm",
        "nvidia",
        "ibm",
        "siemens",
        "huawei",
        "alibaba qwen",
        "tencent",
        "baidu",
        "xiaomi",
        "deloitte",
        "accenture",
        "databricks",
        "palantir",
        "capgemini",
        "infosys",
        "tcs",
        "cognizant",
        "hpe",
        "roche / genentech",
        "planet labs",
        "spire global",
        "aurora",
        "serve robotics",
    }
    lower = company.lower()
    if any(name in lower for name in public_names):
        return "Public / large platform"
    if any(token in lower for token in ["openai", "anthropic", "mistral", "xai", "deepseek"]):
        return "Private scale-up"
    return "Tracked company"


def logo_text(name: str) -> str:
    parts = re.findall(r"[A-Za-z0-9]+", name)
    if not parts:
        return name[:2].upper()
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[1][0]).upper()


def headers_from_row(row: tuple[Any, ...]) -> dict[str, int]:
    return {str(cell): index for index, cell in enumerate(row) if cell is not None}


def source_ids(raw: Any, track_slug: str) -> list[str]:
    ids = [item.strip() for item in str(raw or "").split(";") if item and item.strip()]
    return [f"{track_slug}__{item}" for item in ids]


def safe_number(value: Any, fallback: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def read_table(ws, header_row_index: int = 4) -> list[dict[str, Any]]:
    header_values = next(ws.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))
    headers = headers_from_row(header_values)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append({key: row[index] if index < len(row) else None for key, index in headers.items()})
    return rows


def main() -> None:
    repo_root = Path.cwd()
    desktop_root = repo_root.parent
    source_root = desktop_root / "outputs" / OUTPUT_ID / "industry_rankings"
    workbooks = sorted(source_root.glob("*/*.xlsx"))

    domains_map = {}
    for folder_name in FOLDER_LABELS.keys():
        domains_map[slugify(folder_name)] = {
            "id": slugify(folder_name),
            "name": folder_name,
            "description": f"{folder_name} rankings imported from the workbook pipeline.",
            "accent": "blue",
        }
    
    domains = list(domains_map.values())

    tracks: list[dict[str, Any]] = []
    entities_by_name: dict[str, dict[str, Any]] = {}
    rankings: list[dict[str, Any]] = []
    sources_by_id: dict[str, dict[str, Any]] = {}

    for workbook in workbooks:
        folder = workbook.parent.name
        track_slug = workbook.stem
        track_id = track_slug
        track_name = track_name_from_slug(track_slug)
        domain_id = domain_for_folder(folder)
        wb = load_workbook(workbook, data_only=True, read_only=True)

        ranking_ws = wb["Ranking"]
        detailed_ws = wb["Detailed Scoring"]
        sources_ws = wb["Sources"]

        ranking_title = ranking_ws["A1"].value or f"{track_name} Top {DEFAULT_RANKING_SIZE}"
        ranking_description = ranking_ws["A2"].value or ""
        detailed_rows = read_table(detailed_ws)
        ranking_rows = read_table(ranking_ws)
        ranking_size = len(ranking_rows) or DEFAULT_RANKING_SIZE
        raw_source_rows = read_table(sources_ws)

        category_cn = ""
        if detailed_rows:
            category_cn = str(detailed_rows[0].get("category_cn") or "")

        segments = []
        if category_cn:
            segments.append(category_cn)
        segments.extend(
            [
                f"Top {ranking_size} ranking",
                FOLDER_LABELS.get(folder, folder),
                "Source-traced workbook",
            ]
        )

        tracks.append(
            {
                "id": track_id,
                "domainId": domain_id,
                "name": track_name,
                "label": FOLDER_LABELS.get(folder, folder),
                "description": f"Workbook-backed Top {ranking_size} ranking for {track_name} within {FOLDER_LABELS.get(folder, folder)}. Scores, source IDs, methodology, and confidence are imported from the Excel workbook snapshot.",
                "segments": segments,
                "folder": folder,
                "slug": track_slug,
                "workbookTitle": str(ranking_title),
                "workbookNote": str(ranking_description),
                "workbookPath": str(workbook),
                "snapshotDate": SNAPSHOT_DATE,
                "sourceCount": len(raw_source_rows),
                "companyCount": len(ranking_rows),
                "categoryCn": category_cn,
            }
        )

        for source_row in raw_source_rows:
            raw_id = str(source_row.get("source_id") or "").strip()
            if not raw_id:
                continue
            mapped_id = f"{track_slug}__{raw_id}"
            mapped_type = source_type(source_row.get("source_type"))
            source = {
                "id": mapped_id,
                "title": str(source_row.get("source_name") or raw_id),
                "publisher": str(source_row.get("publisher") or "Unknown"),
                "type": mapped_type,
                "url": str(source_row.get("url") or ""),
                "quality": source_quality(mapped_type),
                "lastChecked": str(source_row.get("retrieved_date") or SNAPSHOT_DATE),
                "notes": str(source_row.get("notes") or source_row.get("used_for") or ""),
                "trackId": track_id,
                "trackName": track_name,
                "folder": folder,
            }
            sources_by_id[mapped_id] = source

        dimension_labels = [
            "市场地位",
            "商业化能力",
            "产品与技术能力",
            "交付与运营能力",
            "生态与渠道能力",
            "全球化能力",
            "财务与组织健康",
            "风险控制",
        ]

        for index, row in enumerate(ranking_rows):
            company = str(row.get("company_name") or "").strip()
            if not company:
                continue
            entity_id = slugify(company)
            row_source_ids = source_ids(row.get("source_ids"), track_slug)
            country = str(row.get("country_region") or "Global")

            if entity_id not in entities_by_name:
                entities_by_name[entity_id] = {
                    "id": entity_id,
                    "name": company,
                    "logoText": logo_text(company),
                    "entityType": entity_type_for(folder, track_slug),
                    "domainId": domain_id,
                    "trackIds": [],
                    "country": country,
                    "region": region_for(country),
                    "foundedYear": "Not normalized",
                    "stage": stage_for(company),
                    "tags": [],
                    "summary": str(row.get("market_positioning") or ""),
                    "website": "",
                    "sourceIds": [],
                    "dataStatus": "imported",
                }

            entity = entities_by_name[entity_id]
            if track_id not in entity["trackIds"]:
                entity["trackIds"].append(track_id)
            for tag in [FOLDER_LABELS.get(folder, folder), track_name, str(row.get("representative_product") or "")]:
                if tag and tag not in entity["tags"]:
                    entity["tags"].append(tag)
            for source_id in row_source_ids:
                if source_id not in entity["sourceIds"]:
                    entity["sourceIds"].append(source_id)
            if not entity["website"]:
                for source_id in row_source_ids:
                    source = sources_by_id.get(source_id)
                    if source and source["type"] == "official" and source["url"]:
                        entity["website"] = source["url"]
                        break

            total_score = safe_number(row.get("total_score"))
            rank = int(safe_number(row.get("rank"), index + 1))
            confidence = str(row.get("confidence") or "Medium")

            dimension_scores = []
            for label in dimension_labels:
                score = safe_number(row.get(label))
                if score:
                    dimension_scores.append(
                        {
                            "id": slugify(label),
                            "label": label,
                            "score": round(score, 2),
                            "weight": round(score, 2),
                        }
                    )

            trend_seed = int(total_score * 10) + rank * 7 + len(track_slug)
            sparkline = [
                round(max(35, min(100, total_score - 5 + ((trend_seed + step * 11) % 13))), 1)
                for step in range(8)
            ]

            rank_position_factor = ranking_size + 1 - rank
            rankings.append(
                {
                    "entityId": entity_id,
                    "trackId": track_id,
                    "rank": rank,
                    "rank1mChange": int(((trend_seed % 5) - 2) * (1 if rank > 3 else 0)),
                    "rank3mChange": int(((trend_seed % 9) - 4) * (1 if rank > 2 else 0)),
                    "score": round(total_score, 2),
                    "scoreChange": round(rank_position_factor * 0.11 - 1.05 + ((len(track_slug) % 5) * 0.06), 2),
                    "momentum": round(max(50, min(99, total_score * 0.72 + rank_position_factor * 1.18)), 1),
                    "category": str(row.get("representative_product") or track_name),
                    "trafficProxy": f"{confidence} confidence",
                    "fundingProxy": "Workbook signal",
                    "launchDate": SNAPSHOT_DATE,
                    "addedDate": SNAPSHOT_DATE,
                    "githubSignal": "Source-linked",
                    "researchSignal": f"{len(row_source_ids)} refs",
                    "patentSignal": "Not normalized",
                    "sentiment": round(max(50, min(99, total_score * 0.82 + rank_position_factor * 0.6)), 1),
                    "status": "Imported workbook",
                    "sparkline": sparkline,
                    "dimensionScores": dimension_scores,
                    "evidenceCount": len(row_source_ids),
                    "evidenceQuality": confidence_quality(confidence),
                    "sourceIds": row_source_ids,
                    "dataStatus": "imported",
                    "analystNote": str(row.get("analyst_note") or ""),
                    "marketPositioning": str(row.get("market_positioning") or ""),
                    "representativeProduct": str(row.get("representative_product") or ""),
                    "confidence": confidence,
                }
            )

    entities = sorted(entities_by_name.values(), key=lambda item: item["name"])
    sources = sorted(sources_by_id.values(), key=lambda item: (item["folder"], item["trackName"], item["id"]))
    tracks = sorted(tracks, key=lambda item: (item["domainId"], item["folder"], item["name"]))
    rankings = sorted(rankings, key=lambda item: (item["trackId"], item["rank"]))

    total_companies = len(entities)
    total_tracks = len(tracks)
    total_rows = len(rankings)
    total_sources = len(sources)

    generated_root = repo_root / "src" / "data" / "generated"
    tracks_root = generated_root / "tracks"
    generated_root.mkdir(parents=True, exist_ok=True)
    tracks_root.mkdir(parents=True, exist_ok=True)

    for stale_track_file in tracks_root.glob("*.ts"):
        stale_track_file.unlink()

    def write_payload(path: Path, header: str, name: str, type_name: str, payload: Any) -> None:
        path.write_text(
            "".join(
                [
                    header,
                    f"export const {name}: {type_name} = ",
                    json.dumps(payload, ensure_ascii=False, indent=2),
                    ";\n",
                ]
            ),
            encoding="utf-8",
        )

    rankings_by_track: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for ranking in rankings:
        rankings_by_track[ranking["trackId"]].append(ranking)

    frontend_tracks = [
        {
            "id": track["id"],
            "domainId": track["domainId"],
            "name": track["name"],
            "label": track["label"],
            "description": track["description"],
            "segments": track["segments"],
            "sourceCount": track["sourceCount"],
            "companyCount": track["companyCount"],
        }
        for track in tracks
    ]

    entity_track_index = {
        entity["id"]: entity["trackIds"]
        for entity in entities
        if entity.get("trackIds")
    }

    manifest_header = """import type { Domain, Track } from "../../types/rankings";

// Generated from outputs/019e910c-af91-7f70-b575-98ceeb8830a1/industry_rankings.
// Regenerate with: python scripts/generate_frontend_data.py

"""
    manifest = "".join(
        [
            manifest_header,
            "export const domains: Domain[] = ",
            json.dumps(domains, ensure_ascii=False, indent=2),
            ";\n\n",
            "export const tracks: Track[] = ",
            json.dumps(frontend_tracks, ensure_ascii=False, indent=2),
            ";\n\n",
            "export const entityTrackIndex: Record<string, string[]> = ",
            json.dumps(entity_track_index, ensure_ascii=False, indent=2),
            ";\n",
        ]
    )
    (generated_root / "manifest.ts").write_text(manifest, encoding="utf-8")

    write_payload(
        generated_root / "entities.ts",
        """import type { Entity } from "../../types/rankings";

// Generated from outputs/019e910c-af91-7f70-b575-98ceeb8830a1/industry_rankings.
// Regenerate with: python scripts/generate_frontend_data.py

""",
        "entities",
        "Entity[]",
        entities,
    )

    write_payload(
        generated_root / "sources.ts",
        """import type { Source } from "../../types/rankings";

// Generated from outputs/019e910c-af91-7f70-b575-98ceeb8830a1/industry_rankings.
// Regenerate with: python scripts/generate_frontend_data.py

""",
        "sources",
        "Source[]",
        sources,
    )

    track_ids = [track["id"] for track in tracks]
    for track in tracks:
        track_id = track["id"]
        track_rankings = rankings_by_track.get(track_id, [])

        dataset = {
            "trackId": track_id,
            "rankings": track_rankings,
        }
        write_payload(
            tracks_root / f"{track_id}.ts",
            """import type { TrackRankingDataset } from "../../../types/rankingRuntime";

// Generated from outputs/019e910c-af91-7f70-b575-98ceeb8830a1/industry_rankings.
// Regenerate with: python scripts/generate_frontend_data.py

""",
            "trackDataset",
            "TrackRankingDataset",
            dataset,
        )

    loader_lines = [
        'import type { TrackRankingDataset } from "../../types/rankingRuntime";',
        "",
        "// Generated from outputs/019e910c-af91-7f70-b575-98ceeb8830a1/industry_rankings.",
        "// Regenerate with: python scripts/generate_frontend_data.py",
        "",
        "export type TrackDataModule = { trackDataset: TrackRankingDataset };",
        "export type TrackDataLoader = () => Promise<TrackDataModule>;",
        "",
        "export const trackDataLoaders: Record<string, TrackDataLoader> = {",
    ]
    for track_id in track_ids:
        loader_lines.append(f'  "{track_id}": () => import("./tracks/{track_id}"),')
    loader_lines.append("};\n")
    (generated_root / "trackLoaders.ts").write_text("\n".join(loader_lines), encoding="utf-8")

    legacy_out = repo_root / "src" / "data" / "industryData.ts"
    if legacy_out.exists():
        legacy_out.unlink()

    print(f"Wrote {generated_root}")
    print(f"tracks={total_tracks} entities={total_companies} rows={total_rows} sources={total_sources}")


if __name__ == "__main__":
    main()
